import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO

def generar_excel_lindo(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    headers = ["Nombre", "Estado", "Hora de reconocimiento"]
    ws.append(headers)

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    for fila in data:
        ws.append(fila)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align

    for column_cells in ws.columns:
        max_length = 0
        col = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
